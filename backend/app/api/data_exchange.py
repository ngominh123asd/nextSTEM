"""Data Exchange API – export & import endpoints for admin panel."""

import csv
import io
import json
import uuid
from datetime import datetime
from xml.etree.ElementTree import Element, SubElement, tostring

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import require_admin
from app.models.user import User, UserPreference
from app.models.admin import AuditLog
from app.models.chat import ChatSession, ChatMessage
from app.services.admin_users import write_audit
from app.utils.security import hash_password

router = APIRouter(prefix="/admin/export", tags=["admin-data"])

# ────────────────────────────────────────────────────────────
# HELPERS
# ────────────────────────────────────────────────────────────

USER_FIELDS = [
    "id", "name", "email", "role", "status", "grade", "school",
    "subjects", "is_active", "created_at", "updated_at",
]


def _serialize(value):
    """Convert a value to a JSON-safe / CSV-safe string."""
    if value is None:
        return ""
    if isinstance(value, (datetime,)):
        return value.isoformat()
    if isinstance(value, (uuid.UUID,)):
        return str(value)
    if isinstance(value, (list,)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, (dict,)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _user_row(u: User) -> dict:
    return {f: _serialize(getattr(u, f, None)) for f in USER_FIELDS}


def _user_dict(u: User) -> dict:
    d = {}
    for f in USER_FIELDS:
        val = getattr(u, f, None)
        if isinstance(val, (datetime,)):
            d[f] = val.isoformat()
        elif isinstance(val, (uuid.UUID,)):
            d[f] = str(val)
        else:
            d[f] = val
    return d


# ────────────────────────────────────────────────────────────
# EXPORT: Users CSV (UTF-8 with BOM for Windows/Excel)
# ────────────────────────────────────────────────────────────
@router.get("/users/csv")
async def export_users_csv(
    admin: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(select(User).order_by(User.created_at.desc()))).scalars().all()

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=USER_FIELDS)
    writer.writeheader()
    for u in rows:
        writer.writerow(_user_row(u))

    # Encode with BOM so Windows/Excel recognises UTF-8
    output = buf.getvalue().encode("utf-8-sig")
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        iter([output]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="users_{now}.csv"'},
    )


# ────────────────────────────────────────────────────────────
# EXPORT: Users Excel (real .xlsx via openpyxl)
# ────────────────────────────────────────────────────────────
@router.get("/users/excel")
async def export_users_excel(
    admin: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(select(User).order_by(User.created_at.desc()))).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1152D4", end_color="1152D4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="D0D0D0"),
    )

    # Write header row
    for col_idx, field in enumerate(USER_FIELDS, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write data rows
    for row_idx, u in enumerate(rows, 2):
        for col_idx, field in enumerate(USER_FIELDS, 1):
            val = getattr(u, field, None)
            if isinstance(val, (datetime,)):
                # Strip timezone – openpyxl doesn't support tz-aware datetimes
                cell = ws.cell(row=row_idx, column=col_idx, value=val.replace(tzinfo=None))
                cell.number_format = "YYYY-MM-DD HH:MM:SS"
            elif isinstance(val, (uuid.UUID,)):
                ws.cell(row=row_idx, column=col_idx, value=str(val))
            elif isinstance(val, (list, dict)):
                ws.cell(row=row_idx, column=col_idx, value=json.dumps(val, ensure_ascii=False))
            elif isinstance(val, bool):
                ws.cell(row=row_idx, column=col_idx, value=val)
            else:
                ws.cell(row=row_idx, column=col_idx, value=val)
            ws.cell(row=row_idx, column=col_idx).border = thin_border

    # Auto-adjust column widths
    for col_idx, field in enumerate(USER_FIELDS, 1):
        max_len = len(field)
        for row_idx in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="users_{now}.xlsx"'},
    )


# ────────────────────────────────────────────────────────────
# EXPORT: Audit Logs JSON
# ────────────────────────────────────────────────────────────
@router.get("/audit-logs/json")
async def export_audit_logs_json(
    admin: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(select(AuditLog).order_by(AuditLog.created_at.desc()))).scalars().all()

    data = []
    for log in rows:
        data.append({
            "id": str(log.id),
            "actor_id": str(log.actor_id) if log.actor_id else None,
            "action": log.action,
            "target_type": log.target_type,
            "target_id": log.target_id,
            "details": log.details,
            "ip_address": log.ip_address,
            "created_at": log.created_at.isoformat() if log.created_at else None,
        })

    content = json.dumps({"audit_logs": data, "count": len(data), "exported_at": datetime.now().isoformat()}, ensure_ascii=False, indent=2)
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        iter([content]),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="audit_logs_{now}.json"'},
    )


# ────────────────────────────────────────────────────────────
# EXPORT: Conversations JSON
# ────────────────────────────────────────────────────────────
@router.get("/conversations/json")
async def export_conversations_json(
    admin: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    sessions = (
        await db.execute(
            select(ChatSession).order_by(ChatSession.created_at.desc())
        )
    ).scalars().all()

    data = []
    for s in sessions:
        messages = (
            await db.execute(
                select(ChatMessage)
                .where(ChatMessage.session_id == s.id)
                .order_by(ChatMessage.created_at.asc())
            )
        ).scalars().all()

        data.append({
            "session_id": str(s.id),
            "user_id": str(s.user_id),
            "mode": s.mode,
            "title": s.title,
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "updated_at": s.updated_at.isoformat() if s.updated_at else None,
            "messages": [
                {
                    "id": str(m.id),
                    "role": m.role,
                    "content": m.content,
                    "created_at": m.created_at.isoformat() if m.created_at else None,
                }
                for m in messages
            ],
        })

    content = json.dumps(
        {"conversations": data, "count": len(data), "exported_at": datetime.now().isoformat()},
        ensure_ascii=False, indent=2,
    )
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        iter([content]),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="conversations_{now}.json"'},
    )


# ────────────────────────────────────────────────────────────
# EXPORT: Full Backup (JSON – all tables)
# ────────────────────────────────────────────────────────────
@router.get("/backup/json")
async def export_full_backup_json(
    admin: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    # Users
    users = (await db.execute(select(User).order_by(User.created_at))).scalars().all()
    users_data = [_user_dict(u) for u in users]

    # User preferences
    prefs = (await db.execute(select(UserPreference))).scalars().all()
    prefs_data = [
        {
            "id": str(p.id), "user_id": str(p.user_id),
            "language": p.language, "notify_email": p.notify_email,
            "notify_push": p.notify_push, "notify_weekly": p.notify_weekly,
        }
        for p in prefs
    ]

    # Audit logs
    logs = (await db.execute(select(AuditLog).order_by(AuditLog.created_at))).scalars().all()
    logs_data = [
        {
            "id": str(l.id), "actor_id": str(l.actor_id) if l.actor_id else None,
            "action": l.action, "target_type": l.target_type, "target_id": l.target_id,
            "details": l.details, "ip_address": l.ip_address,
            "created_at": l.created_at.isoformat() if l.created_at else None,
        }
        for l in logs
    ]

    # Chat sessions + messages
    sessions = (await db.execute(select(ChatSession).order_by(ChatSession.created_at))).scalars().all()
    sessions_data = []
    for s in sessions:
        msgs = (
            await db.execute(
                select(ChatMessage).where(ChatMessage.session_id == s.id).order_by(ChatMessage.created_at)
            )
        ).scalars().all()
        sessions_data.append({
            "id": str(s.id), "user_id": str(s.user_id), "mode": s.mode, "title": s.title,
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "updated_at": s.updated_at.isoformat() if s.updated_at else None,
            "messages": [
                {"id": str(m.id), "role": m.role, "content": m.content, "created_at": m.created_at.isoformat() if m.created_at else None}
                for m in msgs
            ],
        })

    backup = {
        "backup_version": "1.0",
        "exported_at": datetime.now().isoformat(),
        "tables": {
            "users": {"count": len(users_data), "data": users_data},
            "user_preferences": {"count": len(prefs_data), "data": prefs_data},
            "audit_logs": {"count": len(logs_data), "data": logs_data},
            "chat_sessions": {"count": len(sessions_data), "data": sessions_data},
        },
    }

    content = json.dumps(backup, ensure_ascii=False, indent=2)
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        iter([content]),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="nextstem_backup_{now}.json"'},
    )


# ────────────────────────────────────────────────────────────
# EXPORT: Data Dump XML
# ────────────────────────────────────────────────────────────
@router.get("/dump/xml")
async def export_data_dump_xml(
    admin: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    root = Element("nextSTEM_export")
    root.set("exported_at", datetime.now().isoformat())

    # Users
    users = (await db.execute(select(User).order_by(User.created_at))).scalars().all()
    users_el = SubElement(root, "users", count=str(len(users)))
    for u in users:
        ue = SubElement(users_el, "user")
        for f in USER_FIELDS:
            child = SubElement(ue, f)
            child.text = _serialize(getattr(u, f, None))

    # Audit logs
    logs = (await db.execute(select(AuditLog).order_by(AuditLog.created_at))).scalars().all()
    logs_el = SubElement(root, "audit_logs", count=str(len(logs)))
    for l in logs:
        le = SubElement(logs_el, "log")
        for f in ["id", "actor_id", "action", "target_type", "target_id", "ip_address", "created_at"]:
            child = SubElement(le, f)
            child.text = _serialize(getattr(l, f, None))
        det = SubElement(le, "details")
        det.text = json.dumps(l.details, ensure_ascii=False) if l.details else ""

    # Chat sessions
    sessions = (await db.execute(select(ChatSession).order_by(ChatSession.created_at))).scalars().all()
    sess_el = SubElement(root, "chat_sessions", count=str(len(sessions)))
    for s in sessions:
        se = SubElement(sess_el, "session")
        for f in ["id", "user_id", "mode", "title", "created_at", "updated_at"]:
            child = SubElement(se, f)
            child.text = _serialize(getattr(s, f, None))
        msgs = (
            await db.execute(
                select(ChatMessage).where(ChatMessage.session_id == s.id).order_by(ChatMessage.created_at)
            )
        ).scalars().all()
        msgs_el = SubElement(se, "messages", count=str(len(msgs)))
        for m in msgs:
            me = SubElement(msgs_el, "message")
            for mf in ["id", "role", "content", "created_at"]:
                child = SubElement(me, mf)
                child.text = _serialize(getattr(m, mf, None))

    xml_bytes = b'<?xml version="1.0" encoding="UTF-8"?>\n' + tostring(root, encoding="unicode").encode("utf-8")
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        iter([xml_bytes]),
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="nextstem_dump_{now}.xml"'},
    )


# ────────────────────────────────────────────────────────────
# EXPORT: CSV template for import
# ────────────────────────────────────────────────────────────
@router.get("/template/csv")
async def download_csv_template(
    admin: User = Depends(require_admin),
):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["name", "email", "role", "grade", "school"])
    writer.writeheader()
    writer.writerow({"name": "Nguyễn Văn A", "email": "a@example.com", "role": "user", "grade": "12", "school": "THPT ABC"})
    writer.writerow({"name": "Trần Thị B", "email": "b@example.com", "role": "user", "grade": "11", "school": "THPT XYZ"})
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="import_template.csv"'},
    )


# ────────────────────────────────────────────────────────────
# IMPORT: Users from CSV
# ────────────────────────────────────────────────────────────
import_router = APIRouter(prefix="/admin/import", tags=["admin-data"])


@import_router.post("/users/csv")
async def import_users_csv(
    file: UploadFile = File(...),
    admin: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename or not file.filename.endswith(".csv"):
        raise HTTPException(400, "Chỉ chấp nhận file .csv")

    content = await file.read()
    try:
        text = content.decode("utf-8-sig")  # Handle BOM
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    required = {"name", "email"}
    if not required.issubset(set(reader.fieldnames or [])):
        raise HTTPException(400, f"File CSV phải có cột: {', '.join(required)}")

    results = {"created": 0, "skipped": 0, "errors": []}
    default_password = hash_password("ChangeMe@123")

    for i, row in enumerate(reader, start=2):
        email = (row.get("email") or "").strip().lower()
        name = (row.get("name") or "").strip()

        # Validate
        if not email or "@" not in email:
            results["errors"].append({"row": i, "email": email, "reason": "Email không hợp lệ"})
            results["skipped"] += 1
            continue
        if not name or len(name) < 2:
            results["errors"].append({"row": i, "email": email, "reason": "Tên quá ngắn (< 2 ký tự)"})
            results["skipped"] += 1
            continue

        # Check duplicate
        existing = (await db.execute(select(User).where(User.email == email))).scalars().first()
        if existing:
            results["errors"].append({"row": i, "email": email, "reason": "Email đã tồn tại"})
            results["skipped"] += 1
            continue

        role = (row.get("role") or "user").strip().lower()
        if role not in ("user", "admin"):
            role = "user"

        user = User(
            name=name,
            email=email,
            hashed_password=default_password,
            role=role,
            status="active",
            grade=(row.get("grade") or "").strip() or None,
            school=(row.get("school") or "").strip() or None,
        )
        db.add(user)
        results["created"] += 1

    await db.commit()

    await write_audit(
        db, admin.id, "data.import_users",
        details={"created": results["created"], "skipped": results["skipped"]},
    )

    return results


# ────────────────────────────────────────────────────────────
# IMPORT: Preview CSV (parse and validate without creating)
# ────────────────────────────────────────────────────────────
@import_router.post("/users/preview")
async def preview_users_csv(
    file: UploadFile = File(...),
    admin: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename or not file.filename.endswith(".csv"):
        raise HTTPException(400, "Chỉ chấp nhận file .csv")

    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    required = {"name", "email"}
    if not required.issubset(set(reader.fieldnames or [])):
        raise HTTPException(400, f"File CSV phải có cột: {', '.join(required)}")

    preview_rows = []
    errors = []
    for i, row in enumerate(reader, start=2):
        email = (row.get("email") or "").strip().lower()
        name = (row.get("name") or "").strip()
        role = (row.get("role") or "user").strip().lower()
        grade = (row.get("grade") or "").strip()
        school = (row.get("school") or "").strip()

        row_data = {"row": i, "name": name, "email": email, "role": role, "grade": grade, "school": school, "status": "ok"}

        if not email or "@" not in email:
            row_data["status"] = "error"
            row_data["error"] = "Email không hợp lệ"
            errors.append(row_data)
        elif not name or len(name) < 2:
            row_data["status"] = "error"
            row_data["error"] = "Tên quá ngắn"
            errors.append(row_data)
        else:
            existing = (await db.execute(select(User).where(User.email == email))).scalars().first()
            if existing:
                row_data["status"] = "duplicate"
                row_data["error"] = "Email đã tồn tại"

        preview_rows.append(row_data)

        if len(preview_rows) >= 100:  # Limit preview to 100 rows
            break

    valid = sum(1 for r in preview_rows if r["status"] == "ok")
    return {
        "total": len(preview_rows),
        "valid": valid,
        "duplicates": sum(1 for r in preview_rows if r["status"] == "duplicate"),
        "errors": len(errors),
        "rows": preview_rows,
        "columns": list(reader.fieldnames or []),
    }
